"""This module contains the main process of the robot."""

import os
import json
from io import BytesIO
from datetime import datetime
from email.message import EmailMessage
import smtplib

from OpenOrchestrator.orchestrator_connection.connection import OrchestratorConnection
from itk_dev_shared_components.graph import authentication, mail
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
import pyodbc

from robot_framework import config


def process(orchestrator_connection: OrchestratorConnection) -> None:
    """Do the primary process of the robot."""
    orchestrator_connection.log_trace("Running process.")

    excel_file = get_email_attachment(orchestrator_connection)
    wb = read_excel_file(excel_file)

    get_address_changes(wb)
    calculate_difference(wb)

    send_email(wb, orchestrator_connection)

    clear_email_folder(orchestrator_connection)


def get_email_attachment(orchestrator_connection: OrchestratorConnection) -> BytesIO:
    """Find the email in the "AKBO" email folder and get the attached Excel file.

    Args:
        orchestrator_connection: The connection to OpenOrchestrator.

    Raises:
        ValueError: If something doesn't seem right with the email.

    Returns:
        The Excel file as a BytesIO object.
    """
    graph_creds = orchestrator_connection.get_credential(config.GRAPH_API)
    graph_access = authentication.authorize_by_username_password(graph_creds.username, **json.loads(graph_creds.password))

    emails = mail.get_emails_from_folder("itk-rpa@mkb.aarhus.dk", "Indbakke/AKBO", graph_access)
    if len(emails) != 1:
        raise ValueError(f"Unexpected number of emails found: {len(emails)}")

    attachments = mail.list_email_attachments(emails[0], graph_access)
    if len(attachments) != 1:
        raise ValueError(f"Unexpected number of attachments found: {len(attachments)}")

    if not attachments[0].name.endswith(".XLSX"):
        raise ValueError(f"Unexpected attachment found: {attachments[0].name}")

    return mail.get_attachment_data(attachments[0], graph_access)


def read_excel_file(excel_file: BytesIO) -> openpyxl.Workbook:
    """Read the excel sheet and filter away rows that don't meet the criteria.

    Args:
        excel_file: The file to read.

    Returns:
        The Excel workbook after filtering.
    """
    wb = openpyxl.load_workbook(excel_file)
    ws: Worksheet = wb.active

    # Check column names and get indices
    columns = [c.value for c in ws[1]]
    columns.index('CPR')
    columns.index('Beløb')
    udbetalingsdato_index = columns.index('Udbetalingsdato') + 1
    aftaletype_index = columns.index('Aftale type') + 1
    rim_index = columns.index('RIM aftaletype') + 1

    # Iterate rows backwards to delete as we go
    for row_index in range(ws.max_row, 1, -1):
        aftaletype = ws.cell(row_index, aftaletype_index).value
        rim = ws.cell(row_index, rim_index).value
        dato = ws.cell(row_index, udbetalingsdato_index).value
        current_date = datetime.now()

        # Delete where aftaletype is not blank
        # or rim aftaletype is IN
        # or the udbetalingsdato is less than 90 days ago
        if (aftaletype
                or rim == 'IN'
                or (current_date - dato).days < 90):
            ws.delete_rows(row_index)

    return wb


def get_address_changes(wb: openpyxl.Workbook):
    """Find the change date for the address of each person in the Excel file.

    Args:
        wb: The Excel workbook object.
    """
    ws: Worksheet = wb.active
    columns = [c.value for c in ws[1]]

    # Create "Adresseændringsdato" column
    aendringsdato_index = len(columns) + 1
    ws.cell(row=1, column=aendringsdato_index, value="Adresseændringsdato")
    ws.column_dimensions[openpyxl.utils.get_column_letter(aendringsdato_index)].width = 21

    # Get cpr column
    cpr_index = columns.index('CPR') + 1

    # Get dates from database
    conn = pyodbc.connect('DSN=Datavarehuset')
    cursor = conn.cursor()

    # Get change date and write result to Excel file
    for row_index in range(2, ws.max_row+1):
        cpr = ws.cell(row_index, cpr_index).value
        date = look_up_address_change(cursor, cpr)
        if date:
            c = ws.cell(row_index, aendringsdato_index, date)
            c.number_format = 'dd-mm-yyyy'


def look_up_address_change(cursor: pyodbc.Cursor, cpr: str) -> datetime:
    """Look up the last time the given person moved address.
    This is done by look at the address history and finding
    the last time the address actually changed.

    Args:
        cursor: The pyodbc cursor object to use.
        cpr: The cpr of the person to look up.

    Returns:
        _The date of the last address change if any.
    """
    query = "SELECT DatoFra, Adressenoegle FROM DWH.MART.AdresseHistorik WHERE CPR = ? ORDER BY DatoFra DESC"
    cursor.execute(query, cpr)

    if cursor.rowcount == 0:
        return None

    rows = cursor.fetchall()
    date_from = rows[0].DatoFra
    address_key = rows[0].Adressenoegle

    # Iterate through the rows to find the oldest connected occurrence of the address
    for row in rows:
        if row.Adressenoegle == address_key:
            date_from = row.DatoFra
        else:
            break

    return date_from


def calculate_difference(wb: openpyxl.Workbook):
    """Calculate the difference in days from udbetalingsdato to adresseændring.
    If the difference is less than 89 days the row is deleted.

    Args:
        wb: The Excel workbook object.
    """
    ws: Worksheet = wb.active

    # Create new 'Difference' column
    columns = [c.value for c in ws[1]]
    difference_index = len(columns)+1
    ws.cell(row=1, column=difference_index, value="Difference")
    ws.column_dimensions[openpyxl.utils.get_column_letter(difference_index)].width = 12

    # Get other columns
    udbetalingsdato_index = columns.index('Udbetalingsdato')
    aendringsdato_index = columns.index('Adresseændringsdato')

    # Iterate rows backwards to delete as we go
    for row_index in range(ws.max_row, 1, -1):
        row = ws[row_index]
        diff = row[aendringsdato_index].value - row[udbetalingsdato_index].value.date()

        # Insert difference unless diff < 89 then delete row
        if diff.days < 89:
            ws.delete_rows(row_index)
        else:
            ws.cell(row=row_index, column=difference_index, value=diff.days)


def send_email(wb: openpyxl.Workbook, orchestrator_connection: OrchestratorConnection):
    """Send the given workbook in an email to the recipients given as arguments to the OrchestratorConnection.

    Args:
        wb: The excel workbook to send.
        orchestrator_connection: The connection to Orchestrator.
    """
    # Create message
    msg = EmailMessage()
    msg['to'] = orchestrator_connection.process_arguments.split(",")
    msg['from'] = config.EMAIL_SENDER
    msg['subject'] = f"Fastsættelsesliste Hjælp til boligindskud {datetime.now().strftime('%d-%m-%Y')}"
    msg.set_content("Hermed fremsendes den månedlige fastsættelsesliste på Hjælp til boligindskud.")

    # Attach file
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    maintype, subtype = wb.mime_type.split("/")
    file_name = f"Fastsættelsesliste Hjælp til boligindskud {datetime.now().strftime('%d-%m-%Y')}.XLSX"

    msg.add_attachment(excel_file.read(), maintype=maintype, subtype=subtype, filename=file_name)

    # Send message
    with smtplib.SMTP(config.SMTP_SERVER, config.SMTP_PORT) as smtp:
        smtp.starttls()
        smtp.send_message(msg)


def clear_email_folder(orchestrator_connection: OrchestratorConnection):
    """Clear the AKBO email folder

    Args:
        orchestrator_connection: The connection to Orchestrator.
    """
    graph_creds = orchestrator_connection.get_credential(config.GRAPH_API)
    graph_access = authentication.authorize_by_username_password(graph_creds.username, **json.loads(graph_creds.password))

    emails = mail.get_emails_from_folder("itk-rpa@mkb.aarhus.dk", "Indbakke/AKBO", graph_access)
    for email in emails:
        mail.delete_email(email, graph_access)


if __name__ == '__main__':
    conn_string = os.getenv("OpenOrchestratorConnString")
    crypto_key = os.getenv("OpenOrchestratorKey")
    oc = OrchestratorConnection("Akbo Test", conn_string, crypto_key, "ghbm@aarhus.dk,itk-rpa@mkb.aarhus.dk")
    process(oc)
